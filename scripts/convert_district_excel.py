#!/usr/bin/env python3
"""
Convert UP district Excel files into a single JSON for the Santoor dashboard.

Sources:
  - UP_District_Final_Numbers.xlsx  (SER, district, raw census stats, confidence)
  - UP_District_Reach_Model.xlsx    (full reach model with CDMI, reach %, targets)

Output:
  SANTOOR/src/data/up_district_data.json
"""

import json
import os
import math
import openpyxl

BASE = "/Users/prakhar/Desktop/Meegrow/Santoor"
FINAL_NUMBERS = os.path.join(BASE, "UP_District_Final_Numbers.xlsx")
REACH_MODEL = os.path.join(BASE, "UP_District_Reach_Model.xlsx")
OUTPUT = os.path.join(BASE, "SANTOOR", "src", "data", "up_district_data.json")

# Mapping from Excel SER name -> JSON scrKey
SCR_KEY_MAP = {
    "Rohilkhand": "Rohelkhand",
}


def safe_float(val, default=0.0):
    """Convert a cell value to float, returning default for None/NaN."""
    if val is None:
        return default
    try:
        f = float(val)
        return default if math.isnan(f) else f
    except (ValueError, TypeError):
        return default


def safe_str(val, default=""):
    if val is None:
        return default
    return str(val).strip()


def normalize_pct(val, default=0.0):
    """
    Normalize a percentage value to 0-1 range.
    If > 1, assume it is in 0-100 range and divide by 100.
    """
    f = safe_float(val, default)
    if f > 1.0:
        f = f / 100.0
    return round(f, 6)


def main():
    # ── Read Reach Model (primary source — has all columns we need) ──
    wb_reach = openpyxl.load_workbook(REACH_MODEL, data_only=True)
    ws = wb_reach["District Reach Summary"]

    # Column mapping (1-indexed) from header row inspection:
    #  1=SER, 2=District, 3=SYNC Pop, 4=Census Pop (Lakh),
    #  5=Urban%, 6=Literacy%, 7=TV Own%, 8=Electrification%,
    #  9=Mobile%, 10=Internet%, 11=CDMI Score, 12=Alloc% in SER,
    #  13=Pop Share in SER%, 14=Reach%(Overall), 15=Reach%(Urban),
    #  16=Reach%(Rural), 17=Target'000, 18=Reached Pop'000, 19=Confidence Tier

    districts = []
    sers_seen = set()

    for r in range(2, ws.max_row + 1):
        ser = safe_str(ws.cell(r, 1).value)
        district = safe_str(ws.cell(r, 2).value)

        # Skip blank rows, summary rows, and SER summary section
        # (summary rows have numeric "district" values like "21", "9", etc.)
        if not ser or not district or ser == "SER SUMMARY" or ser == "SER":
            continue
        # District names in summary section are just counts (integers)
        try:
            int(district)
            continue  # skip — this is a summary row
        except ValueError:
            pass

        sers_seen.add(ser)
        scr_key = SCR_KEY_MAP.get(ser, ser)

        # SYNC pop is a share of 100Mn — store the raw '000 value from col 17
        sync_pop = safe_float(ws.cell(r, 3).value)          # share (e.g. 0.11439)
        census_pop_lakh = safe_float(ws.cell(r, 4).value)   # e.g. 45.9

        # Percentages — already in 0-1 range in the Reach Model
        urban_pct = normalize_pct(ws.cell(r, 5).value)
        literacy_pct = normalize_pct(ws.cell(r, 6).value)
        tv_pct = normalize_pct(ws.cell(r, 7).value)
        electrification_pct = normalize_pct(ws.cell(r, 8).value)
        mobile_pct = normalize_pct(ws.cell(r, 9).value)
        internet_pct = normalize_pct(ws.cell(r, 10).value)

        cdmi_score = round(safe_float(ws.cell(r, 11).value), 4)
        alloc_pct = normalize_pct(ws.cell(r, 12).value)

        reach_overall = normalize_pct(ws.cell(r, 14).value)
        reach_urban = normalize_pct(ws.cell(r, 15).value)
        reach_rural = normalize_pct(ws.cell(r, 16).value)

        target_pop = safe_float(ws.cell(r, 17).value)       # in '000s
        reached_pop = safe_float(ws.cell(r, 18).value)       # in '000s
        confidence = safe_str(ws.cell(r, 19).value, "Low")

        entry = {
            "ser": ser,
            "scrKey": scr_key,
            "district": district,
            "syncPop": round(sync_pop, 6),
            "censusPop": round(census_pop_lakh, 2),
            "urbanPct": urban_pct,
            "literacyPct": literacy_pct,
            "tvOwnershipPct": tv_pct,
            "mobilePct": mobile_pct,
            "electrificationPct": electrification_pct,
            "internetPct": internet_pct,
            "cdmiScore": cdmi_score,
            "allocPctInSer": alloc_pct,
            "reachOverall": reach_overall,
            "reachUrban": reach_urban,
            "reachRural": reach_rural,
            "targetPop": round(target_pop, 1),
            "reachedPop": round(reached_pop, 1),
            "confidence": confidence,
        }
        districts.append(entry)

    # Canonical SER order
    ser_order = ["Awadh", "Braj", "Bhojpur", "Bundelkhand", "Rohilkhand"]

    output = {
        "metadata": {
            "totalDistricts": len(districts),
            "source": "UP_District_Final_Numbers.xlsx + UP_District_Reach_Model.xlsx",
            "sersFound": sorted(sers_seen),
            "generatedBy": "convert_district_excel.py",
        },
        "sers": ser_order,
        "districts": districts,
    }

    # ── Write JSON ──
    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
    with open(OUTPUT, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    # ── Summary ──
    print(f"Wrote {len(districts)} districts to {OUTPUT}")
    for s in ser_order:
        count = sum(1 for d in districts if d["ser"] == s)
        print(f"  {s}: {count} districts")


if __name__ == "__main__":
    main()
